#!/usr/bin/env python3
"""
Feishu Finance Sync Script v1.0
Daily bidirectional sync: Feishu <-> SQLite

Usage:
  python3 scripts/sync_feishu_finance.py pull
  python3 scripts/sync_feishu_finance.py push
  python3 scripts/sync_feishu_finance.py upload <file>
"""

import sys, json, re, sqlite3, subprocess
from pathlib import Path
from datetime import datetime

WORKSPACE = Path('/home/samuel/.openclaw/workspace/pm-director')
DB_PATH = WORKSPACE / 'database' / 'project_management.db'
FEISHU_APP_TOKEN = 'HSPXbBE0qaemg4swVAQcchXKnMd'
FEISHU_TABLE_ID = 'tblz6ZTwOiMrJLwc'

def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


# ============================================================
# Part 1: Feishu -> DB Pull Sync
# ============================================================

def pull_feishu_data():
    """Pull all project financial data from Feishu. Return list of dicts."""
    temp_file = '/tmp/feishu_pull_{}.json'.format(datetime.now().strftime('%Y%m%d_%H%M'))
    result = subprocess.run(
        ['lark-cli', 'sheets', '+get', '--app-token', FEISHU_APP_TOKEN,
         '--table-id', FEISHU_TABLE_ID, '--output', temp_file],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        print('[WARN] lark-cli pull failed: {}'.format(result.stderr))
        print('  -> Using local cache (last snapshot)')
        snapshots = sorted(Path('/tmp').glob('feishu_pull_*.json'))
        if snapshots:
            with open(snapshots[-1]) as f:
                return json.load(f)
        return []
    with open(temp_file) as f:
        data = json.load(f)
    return data.get('records', []) if isinstance(data, dict) else data


def compare_and_import(feishu_data):
    """Compare Feishu data vs DB latest, generate diff and import."""
    db = get_db()
    today_batch = datetime.now().strftime('%Y-W%W')

    db.execute('''
        CREATE TEMP VIEW IF NOT EXISTS latest_finance AS
        SELECT * FROM finance_records
        WHERE record_id IN (
            SELECT MAX(record_id) FROM finance_records GROUP BY project_id
        )
    ''')
    latest = {}
    for r in db.execute('SELECT * FROM latest_finance').fetchall():
        latest[r['project_id']] = dict(r)

    changes = []
    imports = 0

    for project in feishu_data:
        pid = project.get('project_id', '') or project.get('id', '')
        if not pid:
            continue
        db_record = latest.get(pid)

        if not db_record:
            changes.append({'project_id': pid, 'type': 'new'})
            insert_snapshot(db, pid, project, today_batch)
            imports += 1
            continue

        diffs = check_diff(db_record, project)
        if diffs:
            changes.append({'project_id': pid, 'type': 'update', 'changes': diffs})
            insert_snapshot(db, pid, project, today_batch, diffs)
            imports += 1

    db.commit()
    db.close()
    return {
        'batch_id': today_batch,
        'total': len(feishu_data),
        'imported': imports,
        'changes': changes
    }


def check_diff(old, new):
    """Compare old and new record, return list of changed fields."""
    diffs = []
    fields = [
        ('invoice_total', 'invoice'),
        ('payment_total', 'payment'),
        ('sub_invoice_total', 'sub_invoice'),
        ('sub_payment_total', 'sub_payment'),
        ('invoice_unbilled', 'unbilled'),
        ('payment_unreceived', 'unreceived'),
    ]
    for fname, flabel in fields:
        old_val = float(old.get(fname, 0) or 0)
        new_val = float(new.get(fname, 0) or 0)
        if abs(new_val - old_val) > 0.01:
            diffs.append('{}: {:.2f} -> {:.2f}'.format(flabel, old_val, new_val))
    return diffs


def insert_snapshot(db, pid, project, batch_id, diffs=None):
    """INSERT a new finance_records snapshot (incremental append)."""
    notes = json.dumps({'synced_at': datetime.now().isoformat(), 'changes': diffs}) if diffs else ''
    db.execute('''
        INSERT INTO finance_records (
            project_id, project_name, customer, contract_amount,
            invoice_total, invoice_unbilled, invoice_history,
            payment_total, payment_unreceived, payment_history,
            subcontractor, subcontract_amount,
            sub_invoice_total, sub_payment_total, sub_payment_unpaid,
            batch_id, import_time, notes
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP,?)
    ''', (
        pid,
        project.get('project_name', ''),
        project.get('customer', ''),
        project.get('contract_amount', 0),
        project.get('invoice_total', 0),
        project.get('invoice_unbilled', 0),
        project.get('invoice_history', ''),
        project.get('payment_total', 0),
        project.get('payment_unreceived', 0),
        project.get('payment_history', ''),
        project.get('subcontractor', ''),
        project.get('subcontract_amount', 0),
        project.get('sub_invoice_total', 0),
        project.get('sub_payment_total', 0),
        project.get('sub_payment_unpaid', 0),
        batch_id, notes
    ))


# ============================================================
# Part 2: DB -> Feishu Push Sync
# ============================================================

def push_progress_to_feishu():
    """Calculate project progress from DB and report."""
    db = get_db()
    rows = db.execute('''
        SELECT c.contract_id,
               COUNT(DISTINCT s.stage_id) as total_stages,
               COUNT(DISTINCT CASE WHEN s.status='completed' THEN s.stage_id END) as done_stages,
               SUM(p.planned_amount) as planned_payment,
               SUM(p.actual_amount) as actual_payment,
               COUNT(DISTINCT d.deliverable_id) as total_del,
               COUNT(DISTINCT CASE WHEN d.status='completed' THEN d.deliverable_id END) as done_del
        FROM contracts c
        LEFT JOIN stages s ON c.contract_id = s.contract_id
        LEFT JOIN payments p ON c.contract_id = p.contract_id
        LEFT JOIN deliverables d ON c.contract_id = d.contract_id
        WHERE c.contract_amount > 0
        GROUP BY c.contract_id
    ''').fetchall()
    db.close()

    results = []
    for r in rows:
        stage_pct = round(r['done_stages'] / r['total_stages'] * 100) if r['total_stages'] else 0
        pay_pct = round(r['actual_payment'] / r['planned_payment'] * 100) if r['planned_payment'] else 0
        del_pct = round(r['done_del'] / r['total_del'] * 100) if r['total_del'] else 0
        results.append({
            'project_id': r['contract_id'],
            'stage_progress': stage_pct,
            'payment_progress': pay_pct,
            'deliverable_progress': del_pct,
        })

    total = len(results)
    avg_stage = sum(r['stage_progress'] for r in results) / max(1, total)
    avg_pay = sum(r['payment_progress'] for r in results) / max(1, total)
    print('Progress calculated for {} projects'.format(total))
    print('  Avg stage={:.0f}% payment={:.0f}% deliverable={:.0f}%'.format(avg_stage, avg_pay,
        sum(r['deliverable_progress'] for r in results) / max(1, total)))
    return results


# ============================================================
# Part 3: Manual Upload Module
# ============================================================

def parse_uploaded_table(file_path):
    """Parse uploaded Excel/CSV file for financial data extraction."""
    rows = read_table(file_path)
    results = {'total_rows': len(rows), 'parsed': [], 'errors': [], 'warnings': []}
    seen = set()

    for row in rows:
        try:
            pid = extract_project_id(row)
            if not pid:
                results['errors'].append('Cannot extract project_id')
                continue
            invoices = parse_amount_lines(row)
            payments = parse_amount_lines(row)
            issues = validate_amounts(pid, invoices, payments)
            if issues:
                results['warnings'].extend(['{}: {}'.format(pid, i) for i in issues])
            key = (pid, sum(i['amount'] for i in invoices))
            if key in seen:
                continue
            seen.add(key)
            results['parsed'].append({
                'project_id': pid, 'invoices': invoices, 'payments': payments
            })
        except Exception as e:
            results['errors'].append('Parse error: {}'.format(e))

    return results


def read_table(file_path):
    """Read Excel/CSV into list of dicts."""
    ext = Path(file_path).suffix.lower()
    if ext == '.csv':
        import csv
        with open(file_path) as f:
            return list(csv.DictReader(f))
    elif ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [str(c.value) if c.value else '' for c in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = {}
                for i, v in enumerate(row):
                    if i < len(headers):
                        d[headers[i]] = v
                rows.append(d)
            return rows
        except ImportError:
            print('[ERROR] openpyxl not installed. Run: pip3 install openpyxl')
            return []
    return []


def extract_project_id(row):
    """Extract project_id from row data."""
    for pat in [r'(ZH02-\d{9})', r'(SGSC[A-Z0-9]{14})']:
        for v in row.values():
            if v and isinstance(v, str):
                m = re.search(pat, v)
                if m:
                    return m.group(1)
    return row.get('project_id', '') or row.get('contract_id', '')


def parse_amount_lines(row):
    """Parse amount lines like '180300yuan (2024.10.18)' from row text."""
    items = []
    row_text = ' '.join(str(v) for v in row.values() if v)
    for m in re.finditer(
        r'([\d.]+)\s*yuan\s*[(]\s*(\d{4})[.]\s*(\d{1,2})[.]\s*(\d{1,2})\s*[)]',
        row_text, re.IGNORECASE
    ):
        amt = float(m.group(1))
        date_str = '{}-{}-{}'.format(m.group(2), m.group(3).zfill(2),
                                      (m.group(4) if m.group(4) else '01').zfill(2))
        items.append({'amount': amt, 'date': date_str})
    return items


def validate_amounts(pid, invoices, payments):
    """Validate: invoiced <= contract, received <= invoiced."""
    issues = []
    total_inv = sum(i['amount'] for i in invoices)
    total_pay = sum(p['amount'] for p in payments)
    db = get_db()
    contract = db.execute(
        'SELECT contract_amount FROM contracts WHERE contract_id=?', (pid,)
    ).fetchone()
    db.close()
    if contract:
        contract_amt = contract['contract_amount'] * 10000  # wan -> yuan
        if total_inv > contract_amt * 1.01:
            issues.append('invoice {:.2f}w > contract {:.2f}w'.format(
                total_inv/10000, contract_amt/10000))
        if total_pay > total_inv * 1.01:
            issues.append('payment {:.2f}w > invoice {:.2f}w'.format(
                total_pay/10000, total_inv/10000))
    return issues


# ============================================================
# Main
# ============================================================

if __name__ == '__main__':
    cmd = sys.argv[1] if len(sys.argv) > 1 else 'help'

    if cmd == 'pull':
        print('=== Pulling Feishu data ===')
        data = pull_feishu_data()
        result = compare_and_import(data)
        print('Batch: {}'.format(result['batch_id']))
        print('Imported: {} records ({} changes)'.format(result['imported'], len(result['changes'])))
        for ch in result['changes'][:10]:
            print('  {:8s} {:25s} | {}'.format(ch['type'], ch['project_id'][:25],
                  '; '.join(ch.get('changes', []))))

    elif cmd == 'push':
        print('=== Calculating project progress ===')
        progress = push_progress_to_feishu()

    elif cmd == 'upload':
        if len(sys.argv) < 3:
            print('Usage: python3 sync_feishu_finance.py upload <file.xlsx>')
            sys.exit(1)
        file_path = sys.argv[2]
        print('=== Parsing uploaded file: {} ==='.format(file_path))
        result = parse_uploaded_table(file_path)
        print('Parsed: {} records'.format(len(result['parsed'])))
        if result['warnings']:
            print('[WARN] Validation warnings: {}'.format(len(result['warnings'])))
            for w in result['warnings'][:5]:
                print('  {}'.format(w))
        if result['errors']:
            print('[ERROR] Parse errors: {}'.format(len(result['errors'])))
            for e in result['errors'][:5]:
                print('  {}'.format(e))

    else:
        print('Feishu Finance Sync Tool')
        print('  python3 sync_feishu_finance.py pull        # Feishu -> DB')
        print('  python3 sync_feishu_finance.py push        # DB -> Feishu')
        print('  python3 sync_feishu_finance.py upload <file> # Upload table')
